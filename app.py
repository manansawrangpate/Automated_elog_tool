import streamlit as st
import pandas as pd
import tempfile
import sys
import warnings
import logging
import re
from pathlib import Path
from io import BytesIO

warnings.filterwarnings("ignore")
logging.disable(logging.CRITICAL)
sys.path.insert(0, str(Path(__file__).parent))

st.set_page_config(page_title="Elogbook Sorting Tool")
st.title("Elogbook Sorting Tool")

tab1, tab2, tab3 = st.tabs(["Run triage", "Train model", "Evaluate"])


# ── helpers shared across tabs ─────────────────────────────────────────────

def load_sorted_labels(df: pd.DataFrame, sorted_file) -> pd.DataFrame:
    """Fuzzy-match a SORTED xlsx back onto a RAW DataFrame, setting is_issue=1."""
    from rapidfuzz import fuzz
    from openpyxl import load_workbook

    df = df.copy()
    df["is_issue"] = 0
    df["date_str"] = pd.to_datetime(df["EVENTDATE"], errors="coerce").dt.strftime("%Y-%m-%d")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(sorted_file.read())
        tmp_path = tmp.name

    wb = load_workbook(tmp_path, read_only=True)
    ws = wb["Sheet1"]
    for row in ws.iter_rows(values_only=True):
        if not row[3]:
            continue
        t = str(row[3]).replace("_x000D_", " ").replace("\r\n", " ").replace("\n", " ")
        parts = re.split(r"[\u2014]{5,}", t)
        body = (parts[1].strip() if len(parts) >= 2 else t)[:80]
        date_str = row[2].strftime("%Y-%m-%d") if hasattr(row[2], "strftime") else str(row[2])[:10]
        cands = df[df["date_str"] == date_str]
        best_idx, best_score = None, 0
        for idx, cand in cands.iterrows():
            score = fuzz.partial_ratio(body[:60], str(cand["LDTEXT"] or "").lower()[:80])
            if score > best_score:
                best_score, best_idx = score, idx
        if best_idx is not None and best_score >= 75:
            df.loc[best_idx, "is_issue"] = 1
    return df


# ── TAB 1 — RUN TRIAGE ─────────────────────────────────────────────────────
with tab1:
    st.write("Upload the RAW elogbook export and download the predicted sorted issues.")
    st.info("File must be named **YYYY-MM RAW.xlsx** — e.g. `2026-04 RAW.xlsx`")

    raw_file = st.file_uploader("Choose a RAW elogbook file (.xlsx)", type=["xlsx"], key="run_raw")

    if raw_file:
        st.success("File uploaded successfully!")
        if st.button("Run triage"):
            try:
                import pipeline
                model  = pipeline.load_model()
                prefix = raw_file.name.replace(" RAW.xlsx", "").replace(".xlsx", "")
                df     = pd.read_excel(raw_file)

                if model is None:
                    st.warning("No trained model found — running keyword-only mode. Go to **Train model** tab to train.")

                results  = pipeline.run(df, model=model)
                n_flag   = (results["triage_decision"] == "FLAG").sum()
                n_clear  = (results["triage_decision"] == "CLEAR").sum()
                n_kw     = (results["triage_stage"] == "keyword").sum()
                n_ml     = (results["triage_stage"] == "ml").sum()

                st.write(f"**{n_flag}** predicted issues · **{n_clear}** routine · **{len(results)}** total")
                if model is not None:
                    st.caption(f"Flagged by keyword: {n_kw} · Flagged by ML: {n_ml - (results[(results['triage_stage']=='ml') & (results['triage_decision']=='CLEAR')].shape[0])}")

                st.download_button(
                    label="Download sorted results",
                    data=pipeline.to_excel(results, prefix),
                    file_name=f"{prefix} SORTED predicted.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.exception(e)


# ── TAB 2 — TRAIN MODEL ────────────────────────────────────────────────────
with tab2:
    st.write("Upload one or more RAW + SORTED pairs to train the ML model.")
    st.info(
        "RAW file: **YYYY-MM RAW.xlsx** — e.g. `2026-04 RAW.xlsx`\n\n"
        "SORTED file: **YYYY-MM SORTED.xlsx** — e.g. `2026-04 SORTED.xlsx`"
    )
    st.caption("Upload as many month pairs as you have. More months = better model.")

    train_raws   = st.file_uploader("RAW files (.xlsx)",          type=["xlsx"], key="train_raws",   accept_multiple_files=True)
    train_sorted = st.file_uploader("SORTED answer keys (.xlsx)", type=["xlsx"], key="train_sorted", accept_multiple_files=True)

    if train_raws and train_sorted:
        st.success(f"{len(train_raws)} RAW + {len(train_sorted)} SORTED files uploaded.")

        if st.button("Train model"):
            try:
                import pipeline

                # pair files by matching YYYY-MM prefix
                def get_prefix(name):
                    return name.replace(" RAW.xlsx", "").replace(" SORTED.xlsx", "").replace(".xlsx", "")

                raw_map    = {get_prefix(f.name): f for f in train_raws}
                sorted_map = {get_prefix(f.name): f for f in train_sorted}
                pairs      = [(p, raw_map[p], sorted_map[p])
                              for p in raw_map if p in sorted_map]

                if not pairs:
                    st.error("No matching RAW/SORTED pairs found. Make sure prefixes match (e.g. both start with 2026-04).")
                    st.stop()

                all_texts, all_labels = [], []
                progress = st.progress(0)
                for i, (prefix, raw_f, sorted_f) in enumerate(pairs):
                    df = pd.read_excel(raw_f)
                    df = load_sorted_labels(df, sorted_f)
                    all_texts.extend(df["LDTEXT"].fillna("").tolist())
                    all_labels.extend(df["is_issue"].tolist())
                    progress.progress((i + 1) / len(pairs))

                n_issues  = sum(all_labels)
                n_routine = len(all_labels) - n_issues
                st.write(f"Training on **{len(all_labels):,}** logs · **{n_issues}** issues · **{n_routine:,}** routine")

                with st.spinner("Training..."):
                    model = pipeline.train(all_texts, all_labels)
                    pipeline.save_model(model)

                st.success(f"Model trained on {len(pairs)} month(s) and saved. Switch to **Run triage** to use it.")

            except Exception as e:
                st.exception(e)


# ── TAB 3 — EVALUATE ───────────────────────────────────────────────────────
with tab3:
    st.write("Upload a RAW + SORTED pair to check accuracy.")
    st.info(
        "RAW file: **YYYY-MM RAW.xlsx** — e.g. `2026-04 RAW.xlsx`\n\n"
        "SORTED file: **YYYY-MM SORTED.xlsx** — e.g. `2026-04 SORTED.xlsx`"
    )

    col1, col2 = st.columns(2)
    with col1:
        eval_raw    = st.file_uploader("RAW file (.xlsx)",          type=["xlsx"], key="eval_raw")
    with col2:
        eval_sorted = st.file_uploader("SORTED answer key (.xlsx)", type=["xlsx"], key="eval_sorted")

    if eval_raw and eval_sorted:
        st.success("Both files uploaded!")
        if st.button("Evaluate"):
            try:
                import pipeline
                model = pipeline.load_model()

                df      = pd.read_excel(eval_raw)
                df      = load_sorted_labels(df, eval_sorted)
                results = pipeline.run(df, model=model)

                if model is None:
                    st.warning("No trained model — evaluating keyword-only mode.")

                y_true = df["is_issue"].values
                y_pred = (results["triage_decision"] == "FLAG").astype(int).values

                tp = int(((y_true==1) & (y_pred==1)).sum())
                fp = int(((y_true==0) & (y_pred==1)).sum())
                fn = int(((y_true==1) & (y_pred==0)).sum())
                tn = int(((y_true==0) & (y_pred==0)).sum())
                recall    = round(tp / max(tp+fn, 1), 4)
                precision = round(tp / max(tp+fp, 1), 4)

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Recall",           f"{recall:.4f}")
                c2.metric("Precision",         f"{precision:.4f}")
                c3.metric("Missed (FN)",        fn,
                          delta="CRITICAL" if fn > 0 else "None",
                          delta_color="inverse")
                c4.metric("False alarms (FP)", fp)

                if fn > 0:
                    st.error(f"{fn} real issue(s) missed — add their keywords to KEYWORDS in pipeline.py, then retrain.")
                else:
                    st.success("No issues missed.")

                COLS = [c for c in [
                    "EVENTDATE", "DESCRIPTION", "PERSONGROUP", "LDTEXT",
                    "matched_keyword", "ml_probability", "triage_stage",
                    "is_issue", "triage_decision",
                ] if c in results.columns]

                t1, t2, t3 = st.tabs([
                    f"Caught ({tp})",
                    f"Missed ({fn})",
                    f"False alarms ({fp})",
                ])
                with t1:
                    st.dataframe(results[(y_true==1) & (y_pred==1)][COLS].reset_index(drop=True),
                                 use_container_width=True)
                with t2:
                    if fn > 0:
                        st.dataframe(results[(y_true==1) & (y_pred==0)][COLS].reset_index(drop=True),
                                     use_container_width=True)
                    else:
                        st.info("No missed issues.")
                with t3:
                    st.dataframe(results[(y_true==0) & (y_pred==1)][COLS].head(100).reset_index(drop=True),
                                 use_container_width=True)

                # download report
                buf = BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    pd.DataFrame([{
                        "Total logs": len(df), "True issues": int(y_true.sum()),
                        "Recall": recall, "Precision": precision,
                        "TP": tp, "FP": fp, "FN": fn, "TN": tn,
                    }]).to_excel(writer, sheet_name="Summary", index=False)
                    results[(y_true==1) & (y_pred==0)][COLS].to_excel(writer, sheet_name="MISSED", index=False)
                    results[(y_true==1) & (y_pred==1)][COLS].to_excel(writer, sheet_name="Caught", index=False)
                    results[(y_true==0) & (y_pred==1)][COLS].to_excel(writer, sheet_name="False alarms", index=False)

                prefix = eval_raw.name.replace(" RAW.xlsx", "").replace(".xlsx", "")
                st.download_button(
                    label="Download evaluation report",
                    data=buf.getvalue(),
                    file_name=f"{prefix} evaluation.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            except Exception as e:
                st.exception(e)