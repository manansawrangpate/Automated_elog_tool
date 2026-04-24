"""
data_loader.py

raw csv exports and identified elogs loaded in. normalized and output dataframe ready for training / testing

"""

import re
import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from rapidfuzz import fuzz

logger = logging.getLogger(__name__)


EM_DASH_SEP = re.compile(r"[\u2014]{5,}")
DATE_IN_SORTED = re.compile(r"Date Work Completed:\s*(\d{4}-\d{2}-\d{2})")
TRAILING_META = re.compile(
    r"\s*[•\-]\s*(Attachments|Worked With|REMOTE ENTRY).*$", re.DOTALL
)

RAW_TEXT_COL = "LDTEXT"
RAW_DATE_COL = "EVENTDATE"
RAW_CONTINGENCY_COL = "LOGENTRY_CONTINGENCY"

# fuzzy threshold
FUZZY_MATCH_THRESHOLD = 75

def _normalize_ws(text: str) -> str:
    """collapse whitespace used by operators"""
    if not text:
        return ""
    text = text.replace("\\r\\n", " ").replace("\r\n", " ").replace("\n", " ")
    text = text.replace("_x000D_", " ")
    return re.sub(r"\s+", " ", text).strip()


def _extract_body_from_sorted_entry(full_text: str) -> str:
    """
    strip operator header from sorted entries 
    Strip the operator header from a sorted log entry.

    Input format:
      'Name (Group) @ Site - CODE Date Work Completed: YYYY-MM-DD HH:MM
       ——————————————————————————— LDTEXT BODY ———————————————————————————
       • Attachments: N'

    Returns: just the LDTEXT BODY.
    """
    parts = EM_DASH_SEP.split(full_text)
    if len(parts) >= 2:
        body = parts[1].strip()
        body = TRAILING_META.sub("", body).strip()
        return _normalize_ws(body)
    return _normalize_ws(full_text)


def _extract_date_from_sorted_entry(full_text: str) -> Optional[str]:
    m = DATE_IN_SORTED.search(full_text)
    return m.group(1) if m else None


def _read_sorted_xlsx(path: Path) -> pd.DataFrame:
    """
    pares 1 sorted log into a flat DataFrame. 

    Handles both header formats:
      - With header row:  first cell == 'LOG ID'
      - Without header:   first cell is PERSONGROUP string (e.g. 'WW-CEAST')

    Returns columns: [site_sheet, log_id, date_str, body_text, is_issue]
    """
    wb = load_workbook(str(path), read_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        first_val = all_rows[0][0]
        has_header = isinstance(first_val, str) and first_val.strip() == "LOG ID"
        data_rows = all_rows[1:] if has_header else all_rows

        for row in data_rows:
            if not row or row[0] is None:
                continue

            #if log id header
            if has_header:
                log_id = int(row[0]) if isinstance(row[0], (int, float)) else None
                full_text = row[5] if len(row) > 5 else None
                if not full_text or not isinstance(full_text, str):
                    continue
                date_str = _extract_date_from_sorted_entry(full_text)
                body = _extract_body_from_sorted_entry(full_text)
                is_issue = 1  # everything in the sorted file is an issue
                records.append(
                    {
                        "site_sheet": sheet_name,
                        "log_id": log_id,
                        "date_str": date_str,
                        "body_text": body,
                        "is_issue": is_issue,
                        "confidence": 1.0,
                    }
                )

            #if no log id header
            else:
                # col0=PERSONGROUP, col1=SITE, col2=DATE, col3=LDTEXT,
                # col4=CONFIDENCE, col5=LABEL1, col6=LABEL2
                if len(row) < 4:
                    continue
                date_val = row[2]
                date_str = (
                    date_val.strftime("%Y-%m-%d")
                    if hasattr(date_val, "strftime")
                    else str(date_val)[:10]
                )
                raw_text = row[3]
                if not raw_text or not isinstance(raw_text, str):
                    continue
                body = _normalize_ws(raw_text)
                confidence = float(row[4]) if len(row) > 4 and isinstance(row[4], (int, float)) else 1.0
                label = int(row[5]) if len(row) > 5 and isinstance(row[5], (int, float)) else 1
                records.append(
                    {
                        "site_sheet": sheet_name,
                        "log_id": None,
                        "date_str": date_str,
                        "body_text": body,
                        "is_issue": label,
                        "confidence": confidence,
                    }
                )

    return pd.DataFrame(records)



def _match_sorted_to_raw(
    df_raw: pd.DataFrame, df_sorted: pd.DataFrame
) -> pd.DataFrame:
    """

    Match each sorted issue record to a raw log row
    1. filter raw rows by data
    2. use fuzzy string match on first 6 0chars of body text
    3. if match found (score >= threshold), mark that raw row is_issue=1
    4. unmatch sorted entries are appednmed as synthetic rows with is_issue=1

    returns updated dataframe

    """
    df_raw = df_raw.copy()
    df_raw["is_issue"] = 0
    df_raw["match_confidence"] = 0.0
    df_raw["date_str"] = pd.to_datetime(df_raw[RAW_DATE_COL], errors="coerce").dt.strftime("%Y-%m-%d")
    df_raw["ldtext_norm"] = df_raw[RAW_TEXT_COL].fillna("").apply(_normalize_ws)

    # Always flag LOGENTRY_CONTINGENCY=1 as issues regardless of sorted matching
    contingency_mask = df_raw[RAW_CONTINGENCY_COL] == 1
    df_raw.loc[contingency_mask, "is_issue"] = 1
    df_raw.loc[contingency_mask, "match_confidence"] = 1.0

    unmatched_sorted = []
    matched_count = 0

    for _, sorted_row in df_sorted.iterrows():
        if sorted_row["is_issue"] != 1:
            continue  # skip sorted rows labelled non-issue (rare)

        date = sorted_row["date_str"]
        body = sorted_row["body_text"]
        if not date or not body or len(body) < 10:
            continue

        snippet = body[:60]
        date_mask = df_raw["date_str"] == date
        candidates = df_raw[date_mask]

        best_idx = None
        best_score = 0

        for idx, cand in candidates.iterrows():
            score = fuzz.partial_ratio(snippet, cand["ldtext_norm"][:80])
            if score > best_score:
                best_score = score
                best_idx = idx

        if best_idx is not None and best_score >= FUZZY_MATCH_THRESHOLD:
            df_raw.loc[best_idx, "is_issue"] = 1
            df_raw.loc[best_idx, "match_confidence"] = best_score / 100.0
            matched_count += 1
        else:
            # Unmatched — add as standalone labeled row for training
            unmatched_sorted.append(
                {
                    RAW_TEXT_COL: sorted_row["body_text"],
                    RAW_DATE_COL: sorted_row["date_str"],
                    RAW_CONTINGENCY_COL: 0,
                    "DESCRIPTION": sorted_row["site_sheet"],
                    "PERSONGROUP": "",
                    "LOGBOOK_MONTH": "",
                    "LOCATION": -1,
                    "ISREMOTE": 0,
                    "LOGBOOK_YR": 0,
                    "RELATEDWOTXT": None,
                    "DIRECTEDBY": None,
                    "ELOGBK_DIRECTEDBY": None,
                    "LOGENTRY_OTHEROPS": None,
                    "is_issue": 1,
                    "match_confidence": sorted_row["confidence"],
                    "date_str": sorted_row["date_str"],
                    "ldtext_norm": sorted_row["body_text"],
                }
            )

    logger.info(
        "Matched %d / %d sorted issue records to raw rows (threshold=%d)",
        matched_count,
        len(df_sorted[df_sorted["is_issue"] == 1]),
        FUZZY_MATCH_THRESHOLD,
    )

    if unmatched_sorted:
        df_extra = pd.DataFrame(unmatched_sorted)
        df_raw = pd.concat([df_raw, df_extra], ignore_index=True)

    return df_raw



def load_month_pair(raw_path: str, sorted_path: str) -> pd.DataFrame:
    """
    Load one month's RAW + SORTED pair and return a labeled DataFrame.
    
    raw_path    : path to the RAW CSV export
    sorted_path : path to the SORTED file

    Returns
    DataFrame with all raw columns plus:
        is_issue          : int (0 = routine, 1 = issue)
        match_confidence  : float (0-1, matching confidence)
        date_str          : str  (YYYY-MM-DD)
        ldtext_norm       : str  (normalized LDTEXT)

    """
    raw_path = Path(raw_path)
    sorted_path = Path(sorted_path)

    logger.info("Loading RAW: %s", raw_path.name)
    df_raw = pd.read_csv(str(raw_path), dtype={"LOCATION": str})

    logger.info("Loading SORTED: %s", sorted_path.name)
    df_sorted = _read_sorted_xlsx(sorted_path)

    logger.info(
        "  Raw rows: %d | Sorted issue rows: %d",
        len(df_raw),
        df_sorted["is_issue"].sum(),
    )

    df_labeled = _match_sorted_to_raw(df_raw, df_sorted)

    issue_count = df_labeled["is_issue"].sum()
    logger.info(
        "  Final labeled: %d rows | %d issues (%.1f%%)",
        len(df_labeled),
        issue_count,
        100.0 * issue_count / len(df_labeled),
    )

    return df_labeled


def load_all_pairs(data_dir: str) -> pd.DataFrame:
    """
    discover and load all raw/sorted pairs in a directory
    file names need to be YYYY-MM_RAW.csv  and  YYYY-MM_SORTED.xlsx

    Pairs are matched by their YYYY-MM prefix.
    """
    data_dir = Path(data_dir)
    raw_files = sorted(data_dir.glob("*_RAW.csv"))
    frames = []

    for raw_file in raw_files:
        prefix = raw_file.stem.replace("_RAW", "")
        sorted_file = data_dir / f"{prefix}_SORTED.xlsx"
        if not sorted_file.exists():
            logger.warning("No SORTED file found for %s, skipping", raw_file.name)
            continue
        df = load_month_pair(str(raw_file), str(sorted_file))
        df["source_month"] = prefix
        frames.append(df)

    if not frames:
        raise FileNotFoundError(f"No matching RAW/SORTED pairs found in {data_dir}")

    combined = pd.concat(frames, ignore_index=True)
    logger.info(
        "Loaded %d months | %d total rows | %d issues (%.1f%%)",
        len(frames),
        len(combined),
        combined["is_issue"].sum(),
        100.0 * combined["is_issue"].sum() / len(combined),
    )
    return combined
